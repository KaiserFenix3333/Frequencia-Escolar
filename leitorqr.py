import tkinter as tk
from tkinter import messagebox
import cv2
from pyzbar.pyzbar import decode
from PIL import Image, ImageTk
import os
import openpyxl
from datetime import datetime
import winsound  # Importar módulo para sons (apenas para Windows)

# Adicione as bibliotecas do Google
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# Configuração do Google Drive
SCOPES = ['https://www.googleapis.com/auth/drive.file']
SERVICE_ACCOUNT_FILE = 'client_ico.json'

def upload_to_google_sheets(file_path, file_name):
    # Autenticação
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)

    # Serviço Google Drive
    service = build('drive', 'v3', credentials=creds)

    # Metadados do arquivo para upload
    file_metadata = {
        'name': file_name,
        'mimeType': 'application/vnd.google-apps.spreadsheet'
    }
    media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Faz o upload e converte para o formato Google Planilhas
    file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    print(f'Arquivo enviado e convertido com sucesso. ID do arquivo: {file.get("id")}')

class QRCodeReaderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Leitor de QR Code")
        
        self.camera = cv2.VideoCapture(0)
        self.camera_active = False
        self.student_list = {}  # Alterado para um dicionário
        self.present_students = set()
        
        self.create_widgets()
        self.load_student_list()

    def create_widgets(self):
        # Botão para ligar/desligar o sistema
        self.btn_toggle = tk.Button(self.root, text="Ligar Sistema", command=self.toggle_system)
        self.btn_toggle.pack(pady=10)
        
        # Frame para exibir a imagem da câmera
        self.frame_camera = tk.Label(self.root)
        self.frame_camera.pack()
        
        # Botão para gerar a lista de faltas
        self.btn_generate_absences = tk.Button(self.root, text="Gerar Lista de Faltas", command=self.generate_absences)
        self.btn_generate_absences.pack(pady=10)

    def toggle_system(self):
        if not self.camera_active:
            self.camera_active = True
            self.btn_toggle.config(text="Desligar Sistema")
            self.read_camera()
        else:
            self.camera_active = False
            self.btn_toggle.config(text="Ligar Sistema")
    
    def read_camera(self):
        if self.camera_active:
            ret, frame = self.camera.read()
            if ret:
                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(frame)
                img_tk = ImageTk.PhotoImage(image=img)
                self.frame_camera.img_tk = img_tk
                self.frame_camera.config(image=img_tk)
                
                # Realizar a leitura de QR codes
                decoded_objects = decode(frame)
                if decoded_objects:
                    for obj in decoded_objects:
                        data = obj.data.decode('utf-8')
                        print(f"Dados lidos do QR code: {data}")  # Diagnóstico para ver os dados
                        self.process_qr_code(data)
                        # Emitir o bip sonoro
                        winsound.PlaySound("bip bip.mp3", winsound.SND_FILENAME)
                else:
                    print("Nenhum QR code detectado.")
                
            self.root.after(100, self.read_camera)  # Chamada recursiva para continuar a leitura

    def load_student_list(self):
        # Carregar a lista de alunos do arquivo Excel
        try:
            wb = openpyxl.load_workbook('alunos.xlsx')
            sheet = wb.active
            print("Verificando dados na planilha:")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                print(f"Lendo linha: {row}")
                if row[0] is None:
                    continue
                
                name = (row[0] or '').strip().upper()  # Nome
                serie = (row[1] or '').strip()         # Série
                curso = (row[2] or '').strip()         # Curso
                numero_chamada = (row[3] or '').strip()# Número da Chamada
                
                if name:  # Certifique-se de que o nome não está vazio
                    self.student_list[name] = {
                        'serie': serie,
                        'curso': curso,
                        'numero_chamada': numero_chamada
                    }
            wb.close()
            print(f"Lista de alunos carregada: {self.student_list}")
        except Exception as e:
            print(f"Erro ao carregar a lista de alunos: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao carregar a lista de alunos: {str(e)}")

    def process_qr_code(self, data):
        try:
            # Verificar o formato dos dados
            print(f"Dados recebidos para processamento: {data}")  # Diagnóstico

            # Parse the QR code data
            lines = data.split('\n')
            if len(lines) < 4:
                raise ValueError("Dados insuficientes no QR code. Dados recebidos: " + str(data))

            nome = lines[0].split(': ')[1].strip().upper()
            serie = lines[1].split(': ')[1].strip()
            curso = lines[2].split(': ')[1].strip()
            numero_chamada = lines[3].split(': ')[1].strip()
            
            # Adicionar o aluno à lista de presentes
            self.present_students.add(nome)

            # Definir o caminho para o arquivo Excel
            file_path = "presenca.xlsx"
            
            # Verificar se o arquivo existe, se não, criar um novo
            if not os.path.exists(file_path):
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.append(['Data e Hora', 'Nome', 'Série', 'Curso', 'Número da Chamada'])
                wb.save(file_path)
            
            # Abrir o arquivo Excel e adicionar os dados
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            
            # Adicionar dados na primeira linha disponível
            row = (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), nome, serie, curso, numero_chamada)
            sheet.append(row)
            
            # Salvar as alterações no arquivo Excel
            wb.save(file_path)
            wb.close()
            
            # Mensagem de sucesso
            print(f"Dados de {nome} adicionados à planilha {file_path}")
        
        except Exception as e:
            print(f"Erro ao processar QR code: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao processar QR code: {str(e)}")

    def generate_absences(self):
        # Identificar os alunos faltosos
        absent_students = set(self.student_list.keys()) - self.present_students

        # Definir o caminho para a planilha de faltas
        absences_file_path = "faltas.xlsx"
        
        # Criar a planilha de faltas
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['Nome', 'Série', 'Curso', 'Número da Chamada'])

        # Adicionar os alunos faltosos
        for student in absent_students:
            student_info = self.student_list.get(student, {})
            sheet.append([
                student, 
                student_info.get('serie', ''), 
                student_info.get('curso', ''), 
                student_info.get('numero_chamada', '')
            ])

        wb.save(absences_file_path)
        wb.close()
        
        # Mensagem de sucesso
        print(f"Lista de faltas gerada: {absences_file_path}")
        messagebox.showinfo("Sucesso", f"Lista de faltas gerada: {absences_file_path}")

        # Enviar para o Google Planilhas
        upload_to_google_sheets(absences_file_path, "Lista de Faltas")

if __name__ == "__main__":
    root = tk.Tk()
    app = QRCodeReaderApp(root)
    root.mainloop()
